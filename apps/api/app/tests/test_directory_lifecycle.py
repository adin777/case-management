import io
import json
import uuid
from typing import Self
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.database.session import SessionLocal
from app.main import app
from app.modules.api import password_hash, permissions
from app.modules.approvals.service import create_step_tasks
from app.modules.directory.entra import EntraDirectoryProvider
from app.modules.directory.excel import EXPORT_HEADERS, HEADERS, workbook
from app.modules.directory.fake import FakeDirectoryProvider
from app.modules.directory.sync_service import UserSyncService
from app.modules.environment_assignments.service import apply_rule
from app.modules.models import (
    ApprovalFlowDefinition,
    ApprovalInstance,
    ApprovalStepDefinition,
    ApprovalTask,
    Case,
    Environment,
    EnvironmentAssignmentRule,
    EnvironmentMembership,
    Group,
    GroupMember,
    Permission,
    RequestType,
    Role,
    RolePermission,
    User,
)

client = TestClient(app)

def add_template_row(content: bytes, values: list[str]) -> bytes:
    book = load_workbook(io.BytesIO(content))
    book.active.append(values)
    target_stream = io.BytesIO()
    book.save(target_stream)
    return target_stream.getvalue()


def auth(email: str = "admin@example.com", password: str = "Admin123!") -> dict[str, str]:
    response = client.post("/api/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_manual_user_lifecycle_and_role_no_longer_affects_permissions() -> None:
    headers = auth(); created = client.post("/api/users", headers=headers, json={
        "first_name": "נועה", "last_name": "ישראלי", "display_name": "נועה ישראלי",
        "email": "noa.lifecycle@example.com", "password": "Lifecycle123!", "department": "רכש",
        "job_title": "קניינית", "is_active": True,
    })
    assert created.status_code == 201 and created.json()["source"] == "manual"
    assert created.json()["user_principal_name"] == "noa.lifecycle@example.com"
    user_id = created.json()["id"]
    assert client.patch(f"/api/users/{user_id}", headers=headers, json={"status": "inactive"}).json()["status"] == "inactive"
    assert client.post("/api/auth/login", json={"email": "noa.lifecycle@example.com", "password": "Lifecycle123!"}).status_code == 401
    archived = client.patch(f"/api/users/{user_id}", headers=headers, json={"status": "archived"}).json()
    assert archived["status"] == "archived" and archived["archived_at"]
    with SessionLocal() as db:
        environment = db.scalar(select(Environment)); user = db.get(User, uuid.UUID(user_id)); assert environment and user
        before = permissions(db, user, environment.id)
        role = Role(code=f"legacy_{uuid.uuid4().hex[:8]}", name="Legacy", permissions=["case.lock"])
        db.add(role); db.flush(); permission = db.get(Permission, "case.lock")
        if permission: db.add(RolePermission(role_id=role.id, permission_code=permission.code))
        db.add(EnvironmentMembership(environment_id=environment.id, user_id=user.id, role_id=role.id, source="manual")); db.commit()
        assert permissions(db, user, environment.id) == before


def test_fake_directory_sync_and_manual_inactive_survives() -> None:
    with SessionLocal() as db:
        service = UserSyncService(db, "fake"); batch = FakeDirectoryProvider().fetch_users()
        preview = service.preview(batch); assert preview["created"] == 3
        run = service.apply(batch); db.commit(); assert run.created_count == 3 and run.disabled_count == 0
        dana = db.scalar(select(User).where(User.directory_object_id == "fake-dana")); assert dana and dana.department == "procurement"
        dana.status = "inactive"; dana.is_active = False; db.commit()
        changed = batch.model_copy(deep=True); changed.users[0].job_title = "סמנכ״לית רכש"
        service.apply(changed); db.commit(); db.refresh(dana)
        assert dana.status == "inactive" and dana.is_active is False and dana.job_title == "סמנכ״לית רכש"
        ronit = db.scalar(select(User).where(User.directory_object_id == "fake-ronit")); assert ronit and ronit.status == "inactive"


def test_excel_preview_import_and_export() -> None:
    template = client.get("/api/users/import/template", headers=auth())
    assert template.status_code == 200 and template.content.startswith(b"PK")
    assert [cell.value for cell in next(load_workbook(io.BytesIO(template.content)).active.rows)] == HEADERS
    headers = auth(); content = add_template_row(template.content, ["מאיה", "כהן", "מאיה כהן", "maya.excel@example.com",
        "maya.excel@example.com", "IT", "Help Desk", "03-1", "050-1", "E-1", "PC-1", "True"])
    preview = client.post("/api/users/import/preview", headers=headers,
        files={"file": ("users.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert preview.status_code == 200 and preview.json()["created"] == 1
    applied = client.post("/api/users/import/apply", headers=headers,
        json={"import_session_id": preview.json()["import_session_id"]})
    assert applied.status_code == 200 and applied.json()["created_count"] == 1
    exported = client.get("/api/users-export?source=excel", headers=headers)
    assert exported.status_code == 200 and exported.content.startswith(b"PK")
    exported_book = load_workbook(io.BytesIO(exported.content))
    assert all(field in [cell.value for cell in next(exported_book.active.rows)]
               for field in ["first_name", "last_login_at", "Groups", "Environments"])
    export_columns = {cell.value: cell.column for cell in exported_book.active[1]}
    new_row = exported_book.active.max_row + 1
    for field, value in {"first_name": "נועם", "last_name": "לוי", "display_name": "נועם לוי",
                         "email": "export-roundtrip@example.com", "user_principal_name": "export-roundtrip@example.com",
                         "status": "active", "directory_enabled": "True"}.items():
        exported_book.active.cell(new_row, export_columns[field]).value = value
    exported_stream = io.BytesIO(); exported_book.save(exported_stream)
    export_preview = client.post("/api/users/import/preview", headers=headers,
        files={"file": ("exported-users.xlsx", exported_stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert export_preview.status_code == 200 and export_preview.json()["errors"] == 0
    roundtrip = next(row for row in export_preview.json()["rows"] if row["email"] == "export-roundtrip@example.com")
    assert roundtrip["action"] == "created"
    assert client.post("/api/users/import/apply", headers=headers,
        json={"import_session_id": export_preview.json()["import_session_id"]}).status_code == 200
    with SessionLocal() as db:
        assert db.scalar(select(User).where(User.email == "export-roundtrip@example.com"))
    second_preview = client.post("/api/users/import/preview", headers=headers,
        files={"file": ("users.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert second_preview.json()["created"] == 0


def test_excel_active_markers_are_strict_and_y_creates_active_user() -> None:
    headers = auth()
    template = client.get("/api/users/import/template", headers=headers).content
    for marker, email, expected_active in (
        ("Y", "itay@gmail.com", True),
        ("N", f"inactive-marker-{uuid.uuid4().hex[:6]}@example.com", False),
    ):
        content = add_template_row(template, ["איתי", "נקר", "איתי", email, "יאתי", "IT",
            "בודק תוכנה", "", "", "124" if expected_active else "125", "7709", marker])
        preview = client.post("/api/users/import/preview", headers=headers,
            files={"file": ("users.xlsx", content)})
        assert preview.status_code == 200 and preview.json()["errors"] == 0
        applied = client.post("/api/users/import/apply", headers=headers,
            json={"import_session_id": preview.json()["import_session_id"]})
        assert applied.status_code == 200
        with SessionLocal() as db:
            imported = db.scalar(select(User).where(User.email == email))
            assert imported and imported.is_active is expected_active
        visible = client.get(f"/api/users?active_only={'true' if expected_active else 'false'}", headers=headers)
        assert any(row["email"] == email and row["is_active"] is expected_active for row in visible.json())
    invalid = add_template_row(template, ["Bad", "Marker", "Bad Marker", "bad-marker@example.com",
        "bad-marker@example.com", "", "", "", "", "126", "", "maybe"])
    response = client.post("/api/users/import/preview", headers=headers,
        files={"file": ("users.xlsx", invalid)})
    assert response.status_code == 422 and "Active" in response.json()["detail"]


def test_full_export_snake_case_updates_creates_and_applies_known_memberships() -> None:
    headers = auth()
    with SessionLocal() as db:
        group = db.scalar(select(Group)); environment = db.scalar(select(Environment))
        assert group and environment
        group_name, environment_name = group.name, environment.name_he
    columns = {name:index for index,name in enumerate(EXPORT_HEADERS)}
    def row(email: str, display_name: str, department: str, job_title: str) -> list[str]:
        values = [""] * len(EXPORT_HEADERS)
        values[columns["first_name"]] = display_name
        values[columns["display_name"]] = display_name
        values[columns["email"]] = email
        values[columns["user_principal_name"]] = email
        values[columns["department"]] = department
        values[columns["job_title"]] = job_title
        values[columns["source"]] = "entra"
        values[columns["status"]] = "active"
        values[columns["directory_enabled"]] = "True"
        values[columns["created_at"]] = "1999-01-01T00:00:00Z"
        values[columns["Groups"]] = group_name
        values[columns["Environments"]] = environment_name
        return values
    content = workbook([EXPORT_HEADERS, row("admin@example.com", "מנהל מערכת", "מחלקה מעודכנת", "תפקיד מעודכן"),
                        row("roundtrip-new@example.com", "משתמש חדש", "שירות", "נציג")])
    preview = client.post("/api/users/import/preview", headers=headers, files={"file": ("full-export.xlsx", content)})
    assert preview.status_code == 200, preview.text
    assert preview.json()["updated"] == 1 and preview.json()["created"] == 1
    changed = next(item for item in preview.json()["rows"] if item["email"] == "admin@example.com")
    assert {"department", "job_title"} <= set(changed["changed_fields"])
    applied = client.post("/api/users/import/apply", headers=headers,
        json={"import_session_id": preview.json()["import_session_id"]})
    assert applied.status_code == 200, applied.text
    with SessionLocal() as db:
        admin = db.scalar(select(User).where(User.email == "admin@example.com"))
        created = db.scalar(select(User).where(User.email == "roundtrip-new@example.com"))
        assert admin and admin.department == "מחלקה מעודכנת" and admin.source == "manual"
        assert created and db.get(GroupMember, (group.id, created.id))
        assert db.scalar(select(EnvironmentMembership).where(EnvironmentMembership.environment_id == environment.id,
            EnvironmentMembership.user_id == created.id))


def test_excel_header_diagnostics_name_received_expected_missing_and_extra() -> None:
    template = client.get("/api/users/import/template", headers=auth())
    book = load_workbook(io.BytesIO(add_template_row(template.content, ["Test", "User", "Test User", "test@example.com"])))
    book.active.cell(1, 1).value = "Unexpected Header"
    stream = io.BytesIO(); book.save(stream); content = stream.getvalue()
    response = client.post("/api/users/import/preview", headers=auth(), files={"file": ("bad.xlsx", content)})
    assert response.status_code == 422
    detail = response.json()["detail"]
    assert all(label in detail for label in ("הקובץ אינו בפורמט תקין", "כותרות חסרות", "כותרות לא מוכרות", "Unexpected Header"))


def test_environment_assignment_rule_preserves_manual_membership() -> None:
    with SessionLocal() as db:
        environment = db.scalar(select(Environment)); assert environment
        manual = User(email="manual.rule@example.com", display_name="ידני", password_hash=password_hash.hash("Manual123!"),
            department="IT", job_title="Help Desk", status="active", source="manual", is_active=True)
        generated = User(email="generated.rule@example.com", display_name="כלל", password_hash=password_hash.hash("Rule12345!"),
            department="IT", job_title="Help Desk", status="active", source="manual", is_active=True)
        db.add_all([manual, generated]); db.flush()
        db.add(EnvironmentMembership(environment_id=environment.id, user_id=manual.id, role_id=None, source="manual"))
        rule = EnvironmentAssignmentRule(environment_id=environment.id, name="IT Help Desk",
            conditions_json=[{"field": "department", "value": "IT"}, {"field": "job_title", "value": "Help Desk"}], is_active=True)
        db.add(rule); db.flush(); assert apply_rule(db, rule)["matched"] >= 2; db.commit()
        generated.department = "Finance"; manual.department = "Finance"; result = apply_rule(db, rule); db.commit()
        assert result["removed"] >= 1
        assert db.scalar(select(EnvironmentMembership).where(EnvironmentMembership.user_id == manual.id,
            EnvironmentMembership.source == "manual")) is not None


def test_job_title_approval_first_decision_cancels_snapshot_and_excludes_inactive() -> None:
    with SessionLocal() as db:
        environment = db.scalar(select(Environment)); request_type = db.scalar(select(RequestType));
        reporter = db.scalar(select(User).where(User.email == "admin@example.com")); assert environment and request_type and reporter
        case_item = Case(case_number=f"CASE-JOB-{uuid.uuid4().hex[:8]}", environment_id=environment.id,
            request_type_id=request_type.id, form_definition_id=None, title="Job title approval",
            description="Approval snapshot", reporter_id=reporter.id, requester_id=reporter.id)
        db.add(case_item); db.flush()
        agent = db.scalar(select(User).where(User.email == "agent@example.com")); manager = db.scalar(select(User).where(User.email == "envadmin@example.com")); assert agent and manager
        agent.job_title = manager.job_title = "מאשר רכש"; agent.status = manager.status = "active"; agent.is_active = manager.is_active = True
        for selected in (agent, manager):
            if not db.scalar(select(EnvironmentMembership).where(EnvironmentMembership.environment_id == case_item.environment_id, EnvironmentMembership.user_id == selected.id)):
                db.add(EnvironmentMembership(environment_id=case_item.environment_id, user_id=selected.id, role_id=None, source="manual"))
        flow = ApprovalFlowDefinition(system_number=f"AF-{uuid.uuid4().hex[:8]}", environment_id=case_item.environment_id,
            name="אישור תפקיד", trigger_type="case_created", approval_policy="all_active_steps", is_active=True, created_by=agent.id)
        db.add(flow); db.flush(); step = ApprovalStepDefinition(approval_flow_id=flow.id, step_order=1, name="מאשר רכש",
            approver_type="job_title", approver_job_title="מאשר רכש", required_approvals=1, approval_mode="any", is_active=True)
        db.add(step); db.flush(); instance = ApprovalInstance(system_number=f"AI-{uuid.uuid4().hex[:8]}", case_id=case_item.id,
            approval_flow_id=flow.id, status="pending", current_step_order=1)
        db.add(instance); db.flush(); create_step_tasks(db, instance, 1); instance_id = instance.id; db.commit()
        tasks = list(db.scalars(select(ApprovalTask).where(ApprovalTask.approval_instance_id == instance.id))); assert len(tasks) == 2
        agent_task = next(row for row in tasks if row.approver_user_id == agent.id); task_id = agent_task.id
    decided = client.post(f"/api/approval-tasks/{task_id}/decision", headers=auth("agent@example.com", "Agent123!"), json={"decision": "approved"})
    assert decided.status_code == 200
    with SessionLocal() as db:
        statuses = set(db.scalars(select(ApprovalTask.status).where(ApprovalTask.approval_instance_id == instance_id)))
        assert statuses == {"approved", "cancelled"}


def test_directory_endpoints_contract() -> None:
    headers = auth()
    assert client.get("/api/directory/status", headers=headers).status_code == 200
    diagnostics = client.post("/api/directory/fake/test", headers=headers)
    assert diagnostics.status_code == 200 and diagnostics.json()["ok"] is True
    assert {step["code"] for step in diagnostics.json()["steps"]} == {"provider", "users"}
    preview = client.post("/api/directory/fake/preview", headers=headers)
    assert preview.status_code == 200 and "users" in preview.json()
    applied = client.post("/api/directory/apply", headers=headers, json={"provider": "fake", "users": preview.json()["users"]})
    assert applied.status_code == 200
    assert client.get("/api/directory/runs", headers=headers).status_code == 200


def test_reports_are_real_and_permission_protected() -> None:
    admin_headers=auth(); available=client.get("/api/reports/available",headers=admin_headers)
    assert available.status_code==200 and {row["code"] for row in available.json()}=={"cases","approvals","users","audit"}
    for report in ("approvals","users","audit"):
        response=client.get(f"/api/reports/{report}",headers=admin_headers);assert response.status_code==200 and "items" in response.json()
    assert client.get("/api/reports/users",headers=auth("agent@example.com","Agent123!")).status_code==403


def test_environment_assignment_endpoints_contract() -> None:
    headers = auth()
    environment_id = client.get("/api/environments", headers=headers).json()[0]["id"]
    options = client.get("/api/environment-assignment-options", headers=headers)
    assert options.status_code == 200
    assert {"users", "groups", "departments", "job_titles"} <= options.json().keys()
    payload = {"name": f"IT-{uuid.uuid4().hex[:6]}", "conditions": [{"field": "department", "value": "IT"}], "is_active": True}
    preview = client.post(f"/api/environments/{environment_id}/assignment-rules/preview", headers=headers, json=payload)
    assert preview.status_code == 200 and "matched" in preview.json()
    created = client.post(f"/api/environments/{environment_id}/assignment-rules", headers=headers, json=payload)
    assert created.status_code == 201
    rule_id = created.json()["id"]
    assert client.get(f"/api/environments/{environment_id}/assignment-rules", headers=headers).status_code == 200
    assert client.put(f"/api/environment-assignment-rules/{rule_id}", headers=headers, json={**payload, "is_active": False}).status_code == 200


class _GraphResponse:
    def __init__(self, payload: dict[str, object]): self.payload = payload
    def __enter__(self) -> Self: return self
    def __exit__(self, *_: object) -> None: return None
    def read(self) -> bytes: return json.dumps(self.payload).encode()


def test_entra_provider_mocked_graph_delta_flow() -> None:
    pages = [_GraphResponse({"access_token": "token"}), _GraphResponse({"value": [{"id": "graph-1", "userPrincipalName": "graph@example.com", "displayName": "Graph User", "accountEnabled": True}], "@odata.nextLink": "https://graph/next"}), _GraphResponse({"value": [], "@odata.deltaLink": "https://graph/delta-token"})]
    with patch("app.modules.directory.entra.settings.entra_tenant_id", "tenant"), patch("app.modules.directory.entra.settings.entra_client_id", "client"), patch("app.modules.directory.entra.settings.entra_client_secret", "secret"), patch("app.modules.directory.entra.urllib.request.urlopen", side_effect=pages):
        batch = EntraDirectoryProvider().fetch_users()
    assert batch.delta_link == "https://graph/delta-token"
    assert batch.users[0].directory_object_id == "graph-1"


def test_entra_provider_mocked_connection_diagnostics() -> None:
    responses = [_GraphResponse({"access_token": "token"}), _GraphResponse({"value": [{"id": "graph-1"}]})]
    with patch("app.modules.directory.entra.settings.entra_tenant_id", "tenant"), patch("app.modules.directory.entra.settings.entra_client_id", "client"), patch("app.modules.directory.entra.settings.entra_client_secret", "secret"), patch("app.modules.directory.entra.urllib.request.urlopen", side_effect=responses):
        result = EntraDirectoryProvider().test_connection()
    assert result["ok"] is True
    assert [step["code"] for step in result["steps"]] == ["tenant", "client", "secret", "token", "graph", "users"]
    assert all(step["ok"] for step in result["steps"])
