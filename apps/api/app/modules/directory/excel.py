import io
from collections.abc import Sequence

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from app.modules.directory.provider import NormalizedDirectoryUser

USER_IMPORT_SCHEMA = (("First Name", "first_name"), ("Last Name", "last_name"), ("Display Name", "display_name"),
                      ("Email", "email"), ("Username", "user_principal_name"), ("Department", "department"),
                      ("Job Title", "job_title"), ("Phone", "phone"), ("Mobile Phone", "mobile_phone"),
                      ("Employee ID", "employee_id"), ("Computer Identifier", "computer_identifier"),
                      ("Active", "directory_enabled"))
HEADERS = [column[0] for column in USER_IMPORT_SCHEMA]
FIELDS = [column[1] for column in USER_IMPORT_SCHEMA]
EXPORT_HEADERS = FIELDS[:-1] + ["source", "status", "directory_enabled", "created_at", "updated_at",
                               "last_login_at", "last_directory_sync_at", "Groups", "Environments"]
HEADER_ALIASES = {**{label: field for label, field in USER_IMPORT_SCHEMA}, **{field: field for field in FIELDS},
                  "Groups": "group_names", "Environments": "environment_names", "status": "status"}


def workbook(rows: Sequence[Sequence[object]], sheet_name: str = "Users") -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(list(row))
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def parse(content: bytes) -> list[NormalizedDirectoryUser]:
    try:
        sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
        values = [["" if value is None else str(value) for value in row]
                  for row in sheet.iter_rows(values_only=True)]
    except Exception as exc:
        raise ValueError("הקובץ אינו בפורמט תקין") from exc
    received = values[0] if values else []
    supported = set(HEADERS) | set(EXPORT_HEADERS)
    missing = [header for header in HEADERS if header not in received and HEADER_ALIASES.get(header) not in received]
    extra = [header for header in received if header not in supported]
    if not (set(received) == set(HEADERS) or set(received) == set(EXPORT_HEADERS)):
        raise ValueError("הקובץ אינו בפורמט תקין\nכותרות חסרות:\n- " + ("\n- ".join(missing) or "אין") +
                         "\nכותרות לא מוכרות:\n- " + ("\n- ".join(extra) or "אין"))
    mapped_fields = [HEADER_ALIASES.get(header, header) for header in received]
    result = []
    for number, row in enumerate(values[1:], 2):
        padded = row + [""] * (len(mapped_fields) - len(row))
        data = dict(zip(mapped_fields, padded, strict=True))
        if not data["email"]:
            raise ValueError(f"שורה {number}, Email: התקבל ערך ריק; צפויה כתובת דוא״ל תקינה; שדה חובה")
        data["display_name"] = data["display_name"] or f'{data["first_name"]} {data["last_name"]}'.strip()
        enabled_text = str(data.get("directory_enabled", data.get("status", "active"))).strip().casefold()
        try:
            result.append(NormalizedDirectoryUser(
                first_name=data["first_name"] or None, last_name=data["last_name"] or None,
                display_name=str(data["display_name"]), email=str(data["email"]),
                user_principal_name=data["user_principal_name"] or None, department=data["department"] or None,
                job_title=data["job_title"] or None, phone=data["phone"] or None,
                mobile_phone=data["mobile_phone"] or None, employee_id=data["employee_id"] or None,
                computer_identifier=data["computer_identifier"] or None,
                directory_enabled=enabled_text in {"true", "1", "yes", "כן", "active", "פעיל"},
                group_names=[v.strip() for v in str(data.get("group_names", "")).split(",") if v.strip()],
                environment_names=[v.strip() for v in str(data.get("environment_names", "")).split(",") if v.strip()],
            ))
        except ValidationError as exc:
            field = str(exc.errors()[0].get("loc", ["value"])[0])
            raise ValueError(f"שורה {number}, {field}: התקבל '{data.get(field, '')}'; צפוי פורמט תקין; "
                             f"{exc.errors()[0]['msg']}") from exc
    return result
